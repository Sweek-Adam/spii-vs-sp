"""
SPII vs SP — Version 2 (openpyxl, rapide)
==========================================
Flux entièrement en mémoire, Excel jamais ouvert :
  1. Lecture CSV (avec conversion des décimales à virgule 0,125 -> 0.125)
  2. Construction du modèle métier en mémoire
  3. Appels Jira en parallèle (titres + Story Points + Planning Interval + statut)
  4. Écriture du fichier .xlsx final d'un coup (openpyxl)

Génère un NOUVEAU fichier .xlsx ; l'original n'est pas touché.
Le fichier de sortie peut rester fermé : openpyxl écrit sur disque directement.

Config :  config.toml (équipe, chemins, projet Jira) + secrets.toml (token)
Lance  :  python spii_v2.py
"""

import os
import re
import sys
import time
import subprocess
import unicodedata
from datetime import date, datetime
from concurrent.futures import ThreadPoolExecutor

# tomllib est natif depuis Python 3.11 ; sinon fallback sur le paquet 'tomli'
try:
    import tomllib  # Python >= 3.11
except ModuleNotFoundError:  # Python <= 3.10
    import tomli as tomllib  # pip install tomli

import pandas as pd
import requests
from requests.auth import HTTPBasicAuth

# En entreprise, un proxy d'inspection SSL peut intercepter HTTPS avec son
# propre certificat. truststore fait utiliser à Python le magasin de
# certificats du système (Windows/macOS), qui fait déjà confiance à ce proxy.
# Optionnel : si le paquet n'est pas installé, on continue sans (ex. réseau
# domestique où la vérification standard suffit).
try:
    import truststore
    truststore.inject_into_ssl()
except ImportError:
    pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.chart import PieChart, ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# =====================================================================
# CONFIGURATION
# =====================================================================
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(_BASE_DIR, "config.toml")
SECRETS_PATH = os.path.join(_BASE_DIR, "secrets.toml")

# 12 colonnes de mois : "Mois de référence" = janvier 2026, puis 2/2026 -> 12/2026
MOIS_COLS = ["Mois de référence"] + [f"{m}/2026" for m in range(2, 13)]
# En-têtes affichés dans les onglets : dates 01/01/2026 -> 01/12/2026
ENTETES_MOIS = [date(2026, m, 1) for m in range(1, 13)]

# --- Styles openpyxl ---
FONT_TITRE   = Font(name="Arial", size=16, bold=True, color="003366")
FONT_BOLD    = Font(name="Arial", bold=True)
FILL_ENTETE  = PatternFill("solid", fgColor="E6F0FA")
FILL_TOTAL   = PatternFill("solid", fgColor="F0F0F0")
_thin        = Side(style="thin")
BORDER_ALL   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _lire_toml(chemin):
    """Lit un fichier TOML en tolérant un éventuel BOM en tête de fichier.
    Certains éditeurs / PowerShell écrivent un BOM UTF-8 que tomllib refuse
    (erreur 'Invalid statement' en ligne 1). 'utf-8-sig' le retire au besoin.
    """
    with open(chemin, "r", encoding="utf-8-sig") as f:
        return tomllib.loads(f.read())


def charger_config():
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(f"Config introuvable : {CONFIG_PATH}")
    if not os.path.exists(SECRETS_PATH):
        raise FileNotFoundError(f"Secrets introuvable : {SECRETS_PATH}")
    cfg = _lire_toml(CONFIG_PATH)
    secrets = _lire_toml(SECRETS_PATH)
    cfg.setdefault("jira", {})["api_token"] = \
        secrets.get("jira", {}).get("api_token", "")
    return cfg


# =====================================================================
# 1 & 2. LECTURE CSV + MODÈLE MÉTIER (en mémoire)
# =====================================================================
def conv_num(v):
    """float robuste, gère la virgule décimale française ('0,125' -> 0.125)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", "."))
    except ValueError:
        return 0.0


def _norm(s):
    """Normalise pour comparaison : minuscules, sans accents ni espaces autour.
    'Consommé', ' CONSOMME ', 'consomme' -> 'consomme'
    """
    s = str(s).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def _labels_cat_pourcentage():
    """Étiquettes de camembert affichant la CATÉGORIE + le POURCENTAGE.
    Désactive explicitement nom de série, valeur brute et clé de légende
    (sinon openpyxl affiche 'Série1; Catégorie; Valeur; %').

    Note : les parts à 0 sont exclues en amont (on ne les écrit pas dans le
    tableau source du camembert), donc aucune étiquette "0%" n'apparaît.
    """
    dl = DataLabelList()
    dl.showCatName = True       # ex. "Nom Prénom"
    dl.showPercent = True       # ex. "100%"
    dl.showVal = False          # retire la valeur brute (0,5)
    dl.showSerName = False      # retire "Série1"
    dl.showLegendKey = False
    dl.showBubbleSize = False
    dl.separator = "; "         # séparateur entre catégorie et pourcentage
    return dl


def construire_modele(csv_path, dict_param, prefixe="TCRE"):
    df = pd.read_csv(csv_path, sep=None, engine="python",
                     encoding="latin1", index_col=False)

    # Regex de détection des features : <prefixe>-<numéro> (ex. TCRE-649).
    # re.escape protège un préfixe contenant d'éventuels caractères spéciaux.
    motif_feature = re.compile(rf"{re.escape(prefixe)}-\d{{1,}}", re.IGNORECASE)

    features_conso = {}   # code -> [12 mois]
    collab_data = {}      # nom -> {role, rows, total}
    stats = {}            # code -> ventilation rôle

    for _, row in df.iterrows():
        ressource = str(row["Ressource"]).strip()
        ressource_maj = ressource.upper()
        if ressource_maj not in dict_param:
            continue
        # On ne garde QUE la consommation réelle (exclut Affecte, RAF, Disponible)
        if _norm(row["Type"]) != "consomme":
            continue
        # Exclure les lignes de total agrégé du CSV (Projet == "TOTAL") : elles
        # sont la somme des lignes détaillées du collaborateur. Les inclure
        # provoquerait un double comptage de la consommation.
        if _norm(row["Projet"]) == "total":
            continue
        role = dict_param[ressource_maj].upper()

        livrable = str(row["Livrable"]) if pd.notna(row["Livrable"]) else ""
        match = motif_feature.search(livrable)
        valeurs = [conv_num(row[c]) for c in MOIS_COLS]
        total_ligne = sum(valeurs)

        est_tcre = bool(match)
        code = match.group(0).upper() if est_tcre else ""

        if est_tcre:
            features_conso.setdefault(code, [0.0] * 12)
            for i in range(12):
                features_conso[code][i] += valeurs[i]
            s = stats.setdefault(code, {"total": 0.0, "po_sm": 0.0,
                                        "ba": 0.0, "dev": 0.0, "qa": 0.0})
            s["total"] += total_ligne
            if role in ("PO", "SM"):
                s["po_sm"] += total_ligne
            elif role == "BA":
                s["ba"] += total_ligne
            elif role in ("DEV", "DEVS"):
                s["dev"] += total_ligne
            elif role == "QA":
                s["qa"] += total_ligne

        nom_onglet = ressource[:30]
        c = collab_data.setdefault(nom_onglet, {"role": dict_param[ressource_maj],
                                                "rows": [], "total": 0.0})
        # Les lignes "TOTAL" du CSV ont déjà été exclues plus haut, donc une
        # ligne non-TCRE est forcément une vraie ligne hors feature.
        feat_root = code if est_tcre else "Autre / Hors Feature"
        c["rows"].append([row["Projet"], livrable, feat_root] + valeurs + [total_ligne])
        c["total"] += total_ligne

    return {"features_conso": features_conso,
            "collab_data": collab_data, "stats": stats}


# =====================================================================
# 3. JIRA (parallèle)
# =====================================================================
def recuperer_jira(tcre_list, cfg):
    auth = HTTPBasicAuth(cfg["jira"]["email"], cfg["jira"]["api_token"])
    session = requests.Session()
    session.auth = auth
    url_base = cfg["jira"]["url"]
    sp_field = cfg["jira"]["sp_field"]
    pi_field = cfg["jira"].get("pi_field", "")  # Planning Interval (optionnel)
    projet = cfg["jira"]["projet"]

    def extraire_pi(brut):
        """Le Planning Interval peut être une liste de valeurs. On gère :
        liste d'objets [{'value': 'PI 1'}, ...], liste de chaînes, objet seul
        {'value': 'PI 1'}, ou chaîne simple. Renvoie les valeurs jointes par '; '.
        """
        if brut is None:
            return ""
        if isinstance(brut, list):
            vals = []
            for el in brut:
                if isinstance(el, dict):
                    vals.append(str(el.get("value") or el.get("name") or "").strip())
                else:
                    vals.append(str(el).strip())
            return "; ".join(v for v in vals if v)
        if isinstance(brut, dict):
            return str(brut.get("value") or brut.get("name") or "").strip()
        return str(brut).strip()

    def un_tcre(tcre):
        titre = "Titre introuvable"
        pi = ""
        statut = ""
        # Champs demandés au ticket : titre + statut + Planning Interval (si configuré)
        champs = "summary,status" + (f",{pi_field}" if pi_field else "")
        try:
            r = session.get(f"{url_base}/rest/api/3/issue/{tcre}",
                            params={"fields": champs}, timeout=30)
            if r.status_code == 200:
                fields = r.json().get("fields", {})
                titre = fields.get("summary", "Sans titre")
                statut = (fields.get("status") or {}).get("name", "")
                if pi_field:
                    pi = extraire_pi(fields.get(pi_field))
        except Exception as e:
            print(f"   Erreur titre/PI {tcre} : {e}")
        jql = f'parent in ("{tcre}") AND PROJECT = "{projet}"'
        total_sp = 0.0
        try:
            r = session.get(f"{url_base}/rest/api/3/search/jql",
                            params={"jql": jql, "fields": sp_field,
                                    "maxResults": 100}, timeout=30)
            if r.status_code == 200:
                for issue in r.json().get("issues", []):
                    total_sp += float(issue["fields"].get(sp_field) or 0)
        except Exception as e:
            print(f"   Erreur SP {tcre} : {e}")
        return tcre, {"titre": titre, "sp": total_sp, "pi": pi, "statut": statut}

    with ThreadPoolExecutor(max_workers=8) as ex:
        return dict(ex.map(un_tcre, tcre_list))


# =====================================================================
# 4. ÉCRITURE OPENPYXL
# =====================================================================
def _style_entete(ws, cell_range):
    for row in ws[cell_range]:
        for c in row:
            c.font = FONT_BOLD
            c.fill = FILL_ENTETE


def _bordures(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER_ALL


def _couleur_degrade(t):
    """Interpole vert (t=0) -> jaune (t=0.5) -> rouge (t=1). Renvoie un hex 'RRGGBB'."""
    t = max(0.0, min(1.0, t))
    vert, jaune, rouge = (99, 190, 123), (255, 235, 132), (248, 105, 107)
    if t <= 0.5:
        f, a, b = t / 0.5, vert, jaune
    else:
        f, a, b = (t - 0.5) / 0.5, jaune, rouge
    rgb = tuple(int(round(a[i] + (b[i] - a[i]) * f)) for i in range(3))
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


def _appliquer_degrade(ws, col_letter, row_min, row_max, lignes_exclues=None):
    """Colore le fond des cellules d'une colonne selon leur valeur (vert->rouge).
    Couleur figée (cohérent avec le choix valeurs figées de la V2).

    lignes_exclues : ensemble de numéros de lignes (absolus) à NE PAS colorer.
    Ces lignes sont aussi ignorées dans le calcul du min/max, pour ne pas
    fausser l'échelle du dégradé.
    """
    lignes_exclues = lignes_exclues or set()
    vals = []
    for r in range(row_min, row_max + 1):
        if r in lignes_exclues:
            vals.append(None)  # ni colorée, ni comptée dans l'échelle
            continue
        v = ws[f"{col_letter}{r}"].value
        vals.append(v if isinstance(v, (int, float)) else None)
    nums = [v for v in vals if v is not None]
    if not nums:
        return
    vmin, vmax = min(nums), max(nums)
    etendue = (vmax - vmin) or 1.0
    for i, v in enumerate(vals):
        if v is not None:
            hexc = _couleur_degrade((v - vmin) / etendue)
            ws[f"{col_letter}{row_min + i}"].fill = PatternFill("solid", fgColor=hexc)


def _ajuster_colonnes(ws, largeur_min=8, largeur_max=45):
    """Ajuste la largeur de chaque colonne au contenu le plus long.

    openpyxl ne sait pas mesurer le rendu réel du texte : on approxime par le
    nombre de caractères. Les bornes évitent des colonnes trop étroites ou,
    à l'inverse, démesurées (ex. un titre TCRE très long).

    Ignore les cellules fusionnées et les graphiques (qui ne portent pas de
    largeur de colonne).
    """
    largeurs = {}  # lettre de colonne -> longueur max observée
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            # Longueur du texte affiché. Pour les dates, on estime ~10 (jj/mm/aaaa).
            if hasattr(cell.value, "strftime"):
                longueur = 10
            else:
                # Pour un nombre à 3 décimales, la repr str suffit comme approximation
                longueur = len(str(cell.value))
            col = cell.column_letter
            if longueur > largeurs.get(col, 0):
                largeurs[col] = longueur

    for col, longueur in largeurs.items():
        # +2 de marge pour respirer ; borné entre min et max
        largeur = max(largeur_min, min(longueur + 2, largeur_max))
        ws.column_dimensions[col].width = largeur


def _ajouter_legende_degrade(ws, cell_ancre):
    """Ajoute une petite légende expliquant le dégradé de couleurs.
    Vert = valeur basse, Rouge = valeur élevée.
    `cell_ancre` est la cellule de départ (ex. "A15"), la légende occupe 3 lignes.
    """
    from openpyxl.utils.cell import coordinate_to_tuple
    r0, c0 = coordinate_to_tuple(cell_ancre)  # (ligne, colonne)
    # Titre
    titre = ws.cell(row=r0, column=c0, value="Légende couleurs :")
    titre.font = FONT_BOLD
    # 3 paliers : vert (bas), jaune (moyen), rouge (élevé)
    paliers = [
        (_couleur_degrade(0.0), "Consommation basse"),
        (_couleur_degrade(0.5), "Consommation moyenne"),
        (_couleur_degrade(1.0), "Consommation élevée"),
    ]
    for i, (hexc, libelle) in enumerate(paliers):
        rr = r0 + 1 + i
        case_couleur = ws.cell(row=rr, column=c0)
        case_couleur.fill = PatternFill("solid", fgColor=hexc)
        case_couleur.border = BORDER_ALL
        ws.cell(row=rr, column=c0 + 1, value=libelle)


def ecrire_classeur(modele, jira, sortie_path, cfg):
    wb = Workbook()
    wb.remove(wb.active)  # retire la feuille vide par défaut

    features = modele["features_conso"]
    collab = modele["collab_data"]
    stats = modele["stats"]

    # Préfixe des features (ex. "TCRE"), configurable.
    prefixe = str(cfg["jira"].get("prefixe_feature", "TCRE")).strip() or "TCRE"

    # Tri des features par numéro
    motif_num = re.compile(rf"{re.escape(prefixe)}-(\d+)", re.IGNORECASE)
    def num(code):
        m = motif_num.search(code)
        return int(m.group(1)) if m else 0
    codes_tries = sorted(features.keys(), key=num)

    # TCRE qui auront un onglet dédié (conso > 0) : sert aux liens internes.
    codes_avec_onglet = {c for c in codes_tries if stats[c]["total"] > 0}

    # Style de lien interne (bleu souligné, comme un hyperlien classique)
    FONT_LIEN = Font(name="Arial", color="0563C1", underline="single")

    projet = cfg["jira"]["projet"]
    # --- Suivi_Features ---
    ws_feat = wb.create_sheet("Suivi_Features_" + projet)
    ws_feat.append(["Code Feature", "Planning Interval", "Statut"]
                   + ENTETES_MOIS + ["Total Consommé"])
    for code in codes_tries:
        mois = features[code]
        info = jira.get(code, {})
        pi = info.get("pi", "")
        statut = info.get("statut", "")
        ws_feat.append([code, pi, statut] + mois + [sum(mois)])
    _style_entete(ws_feat, "A1:P1")
    # Dégradé vert->rouge sur la colonne Total Consommé (P, décalée par PI+Statut)
    if codes_tries:
        _appliquer_degrade(ws_feat, "P", 2, 1 + len(codes_tries))
    # Légende des couleurs du dégradé (à droite du tableau, colonne R)
    _ajouter_legende_degrade(ws_feat, "R2")

    # --- Onglets collaborateurs ---
    for nom, data in collab.items():
        ws = wb.create_sheet(nom[:31])  # limite Excel : 31 car.
        ws.append(["Projet", "Livrable d'origine", "Feature (Racine)",
                   "Planning Interval"] + ENTETES_MOIS + ["Total"])
        for r in data["rows"]:
            # r = [Projet, Livrable, Feature, 12 mois, Total]
            # On insère le Planning Interval (du TCRE) juste après la Feature.
            code_feat = str(r[2]).upper()
            pi = jira.get(code_feat, {}).get("pi", "")
            ws.append(r[:3] + [pi] + r[3:])
        _style_entete(ws, "A1:Q1")  # une colonne de plus (Q au lieu de P)
        # Rôle en S1 (colonne 19, décalé de +1)
        ws.cell(row=1, column=18, value="Rôle :").font = FONT_BOLD
        ws.cell(row=1, column=19, value=data["role"])
        # Dégradé vert->rouge sur la colonne Total (Q maintenant) — lignes détail.
        # On exclut les lignes "Indispo DOSI ACCORDS".
        n_detail = len(data["rows"])
        if n_detail >= 1:
            lignes_indispo = {
                2 + i for i, r in enumerate(data["rows"])
                if str(r[0]).strip() == "Indispo DOSI ACCORDS"
            }
            _appliquer_degrade(ws, "Q", 2, 1 + n_detail,
                               lignes_exclues=lignes_indispo)

        # Ligne TOTAL globale calculée (somme des lignes détail), affichage seul.
        if n_detail >= 1:
            row_total = 2 + n_detail
            # Somme colonne par colonne : 12 mois + Total (cols E..Q = idx 4..16)
            sommes = [0.0] * 13  # 12 mois + total
            for r in data["rows"]:
                for j in range(13):           # r[3..15] = 12 mois + total
                    sommes[j] += conv_num(r[3 + j])
            ws.cell(row=row_total, column=1, value="TOTAL").font = FONT_BOLD
            # Les valeurs commencent en colonne E (5) à cause de la col PI insérée
            for j, val in enumerate(sommes):
                c = ws.cell(row=row_total, column=5 + j, value=round(val, 3))
                c.font = FONT_BOLD
            for col in range(1, 18):          # fond gris clair sur A..Q
                ws.cell(row=row_total, column=col).fill = FILL_TOTAL

        # Légende des couleurs du dégradé (colonne S, sous le "Rôle")
        _ajouter_legende_degrade(ws, "S3")

    # --- Stats ---
    ws_stats = wb.create_sheet("Stats")
    # Ordre : Feature, Statut, Planning Interval, puis les indicateurs.
    ws_stats.append(["Feature", "Statut", "Planning Interval",
                     "Story points", "Total consommé",
                     "Ratio total consommé / story points",
                     "Conso PO / SM", "Conso BA", "Conso Dévs", "Conso QA",
                     "Titre", "Lien Jira"])
    _style_entete(ws_stats, "A1:L1")
    r = 2
    # URL de base Jira (pour les liens cliquables vers chaque feature)
    url_base_jira = str(cfg["jira"].get("url", "")).rstrip("/")
    # Lignes de données + collecte (sp, total) par feature pour la synthèse
    sp_total_par_feature = []  # [(sp, total), ...]
    for code in codes_tries:
        s = stats[code]
        sp = jira.get(code, {}).get("sp", 0.0)
        titre = jira.get(code, {}).get("titre", "")
        pi = jira.get(code, {}).get("pi", "")
        statut = jira.get(code, {}).get("statut", "")
        # Ratio arrondi à 3 décimales en amont : combiné au format "General",
        # un entier s'affiche sans virgule (5 -> "5") et un décimal garde au
        # plus 3 chiffres (14,285714... -> 14,286).
        ratio = round(s["total"] / sp, 3) if sp else 0
        ws_stats.append([code, statut, pi, sp, s["total"], ratio,
                         s["po_sm"], s["ba"], s["dev"], s["qa"], titre])
        # Ratio en colonne F : format General (pas de virgule parasite).
        ws_stats.cell(row=r, column=6).number_format = "General"
        # Lien interne vers l'onglet TCRE, seulement s'il existe (conso > 0)
        if code in codes_avec_onglet:
            cell = ws_stats.cell(row=r, column=1)
            cell.hyperlink = f"#'{code}'!A1"
            cell.font = FONT_LIEN
        # Lien cliquable vers la page Jira de la feature (colonne L)
        if url_base_jira:
            cell_jira = ws_stats.cell(row=r, column=12, value="Ouvrir ↗")
            cell_jira.hyperlink = f"{url_base_jira}/browse/{code}"
            cell_jira.font = FONT_LIEN
        sp_total_par_feature.append((sp, s["total"]))
        r += 1
    derniere = r - 1

    # Ligne TOTAL — somme des colonnes numériques : Story points (D), Total (E),
    # Conso PO/SM (G), BA (H), Dévs (I), QA (J).
    ws_stats.cell(row=r, column=1, value="TOTAL").font = FONT_BOLD
    for col in (4, 5, 7, 8, 9, 10):
        L = get_column_letter(col)
        ws_stats.cell(row=r, column=col, value=f"=SUM({L}2:{L}{derniere})").font = FONT_BOLD
    for c in range(1, 11):
        ws_stats.cell(row=r, column=c).fill = FILL_TOTAL
    total_row = r

    # Ligne MOYENNE J/SP TYPOLOGIE (moyenne globale = total conso / total SP)
    avg_row = total_row + 1
    somme_sp = sum(sp for sp, _ in sp_total_par_feature)
    somme_total = sum(t for _, t in sp_total_par_feature)
    moy_globale = (somme_total / somme_sp) if somme_sp else 0.0
    ws_stats.cell(row=avg_row, column=1, value="MOYENNE J/SP TYPOLOGIE").font = FONT_BOLD
    # La moyenne s'affiche dans la colonne Ratio (F).
    cell_moy = ws_stats.cell(row=avg_row, column=6, value=round(moy_globale, 3))
    cell_moy.number_format = '0.000" jour(s)"'
    for c in range(1, 11):
        ws_stats.cell(row=avg_row, column=c).fill = PatternFill("solid", fgColor="E1EBF5")

    # Dégradé vert->rouge sur la colonne Total consommé (E), lignes de données
    # uniquement (on exclut les lignes TOTAL et MOYENNE en dessous).
    if derniere >= 2:
        _appliquer_degrade(ws_stats, "E", 2, derniere)
    # Légende des couleurs (à droite, colonne N — A..L occupées par le tableau).
    _ajouter_legende_degrade(ws_stats, "N2")

    # --- Tableau de synthèse : moyenne jours réels par complexité (SP) ---
    start_recap = derniere + 5
    ws_stats.cell(row=start_recap, column=1, value="Complexité (SP)")
    ws_stats.cell(row=start_recap, column=2, value="Moyenne Jours Réels")
    ws_stats.cell(row=start_recap, column=3, value="Nb Features")
    for c in range(1, 4):
        ws_stats.cell(row=start_recap, column=c).font = FONT_BOLD
        ws_stats.cell(row=start_recap, column=c).fill = PatternFill("solid", fgColor="D2E1F0")

    # Regroupement par SP unique (> 0), calculé en Python
    par_sp = {}  # sp -> [liste des totaux]
    for sp, total in sp_total_par_feature:
        if isinstance(sp, (int, float)) and sp > 0:
            par_sp.setdefault(sp, []).append(total)

    recap_first = start_recap + 1
    for i, sp_val in enumerate(sorted(par_sp.keys())):
        totaux = par_sp[sp_val]
        moyenne = sum(totaux) / len(totaux)
        rr = recap_first + i
        ws_stats.cell(row=rr, column=1, value=sp_val)
        ws_stats.cell(row=rr, column=2, value=round(moyenne, 2)).number_format = '0.00'
        ws_stats.cell(row=rr, column=3, value=len(totaux))
    recap_last = recap_first + len(par_sp) - 1

    # --- Graphique nuage de points : SP (X) vs Total consommé (Y), 1 point/TCRE ---
    if derniere >= 2:
        scatter = ScatterChart()
        scatter.title = f"Features {prefixe} : Story Points vs Jours réels"
        scatter.x_axis.title = "Story Points"
        scatter.y_axis.title = "Total consommé (jours)"
        scatter.height, scatter.width = 9, 14
        # Graduations : marques majeures vers l'extérieur + quadrillage majeur,
        # et on force l'affichage des étiquettes de graduation sur les 2 axes.
        scatter.x_axis.majorTickMark = "out"
        scatter.y_axis.majorTickMark = "out"
        scatter.x_axis.minorTickMark = "out"
        scatter.y_axis.minorTickMark = "out"
        scatter.x_axis.majorGridlines = ChartLines()
        scatter.y_axis.majorGridlines = ChartLines()
        scatter.x_axis.tickLblPos = "low"
        scatter.y_axis.tickLblPos = "nextTo"
        scatter.x_axis.delete = False  # ne pas masquer l'axe (défaut openpyxl)
        scatter.y_axis.delete = False
        xref = Reference(ws_stats, min_col=4, min_row=2, max_row=derniere)  # SP (col D)
        yref = Reference(ws_stats, min_col=5, min_row=2, max_row=derniere)  # Total (col E)
        serie = Series(yref, xref, title=f"Features {prefixe}")
        serie.marker.symbol = "circle"
        serie.graphicalProperties.line.noFill = True  # points seuls, pas de ligne
        scatter.series.append(serie)
        ws_stats.add_chart(scatter, f"E{start_recap}")

    # --- Onglets par TCRE (avec camemberts) ---
    # On ne crée un onglet dédié que si le TCRE a une consommation > 0.
    # (Les lignes restent présentes dans Stats et Suivi_Features.)
    for idx, code in enumerate(codes_tries):
        s = stats[code]
        if s["total"] <= 0:
            continue
        titre = jira.get(code, {}).get("titre", "")
        ws = wb.create_sheet(code)

        # Titre
        ws.cell(row=1, column=1, value=f"{code} - {titre}").font = FONT_TITRE
        ws.row_dimensions[1].height = 30

        # Planning Interval (ligne 2, entre le titre et le lien retour)
        pi = jira.get(code, {}).get("pi", "")
        ws.cell(row=2, column=1, value="Planning Interval :").font = FONT_BOLD
        ws.cell(row=2, column=2, value=pi if pi else "—")

        # Lien retour vers l'onglet Stats (ligne 3, colonne A)
        lien = ws.cell(row=3, column=1, value="← Retour vers Stats")
        lien.hyperlink = "#'Stats'!A1"
        lien.font = FONT_LIEN

        # Lien vers la page Jira du TCRE (ligne 3, colonne B, à droite du retour)
        url_base = str(cfg["jira"].get("url", "")).rstrip("/")
        if url_base:
            lien_jira = ws.cell(row=3, column=2, value="Ouvrir dans Jira ↗")
            lien_jira.hyperlink = f"{url_base}/browse/{code}"
            lien_jira.font = FONT_LIEN

        # Tableau 1 : par profil (affiché en entier, y compris les profils à 0)
        ws.cell(row=4, column=1, value="Profil / Rôle")
        ws.cell(row=4, column=2, value="Consommation (Jours)")
        ws.cell(row=4, column=3, value="%")
        profils = [("PO / SM", s["po_sm"]), ("BA", s["ba"]),
                   ("Développeurs", s["dev"]), ("QA", s["qa"])]
        total_profils = sum(v for _, v in profils)
        for i, (lab, val) in enumerate(profils):
            ws.cell(row=5 + i, column=1, value=lab)
            ws.cell(row=5 + i, column=2, value=round(val, 3))
            # Pourcentage de la ligne dans le total (vide si total nul)
            pct = ws.cell(row=5 + i, column=3,
                          value=(val / total_profils) if total_profils else 0)
            pct.number_format = "0.0%"
        ws.cell(row=9, column=1, value="TOTAL")
        ws.cell(row=9, column=2, value=round(total_profils, 3))
        pct_tot = ws.cell(row=9, column=3, value=1 if total_profils else 0)
        pct_tot.number_format = "0.0%"
        _style_entete(ws, "A4:C4")
        ws.cell(row=9, column=1).font = FONT_BOLD
        ws.cell(row=9, column=2).font = FONT_BOLD
        ws.cell(row=9, column=3).font = FONT_BOLD
        ws.cell(row=9, column=1).fill = FILL_TOTAL
        ws.cell(row=9, column=2).fill = FILL_TOTAL
        ws.cell(row=9, column=3).fill = FILL_TOTAL
        _bordures(ws, 4, 1, 9, 3)

        # Camembert 1 : par profil — sur les profils NON NULS uniquement.
        # On écrit les parts > 0 dans une zone source dédiée (colonnes V/W,
        # à l'écart) pour que le camembert n'affiche ni tranche ni "0%".
        profils_non_nuls = [(lab, val) for lab, val in profils if val > 0]
        if profils_non_nuls:
            COL_LAB, COL_VAL = 22, 23   # V et W
            for i, (lab, val) in enumerate(profils_non_nuls):
                ws.cell(row=5 + i, column=COL_LAB, value=lab)
                ws.cell(row=5 + i, column=COL_VAL, value=round(val, 3))
            r_fin = 5 + len(profils_non_nuls) - 1
            pie1 = PieChart()
            pie1.title = "Répartition par Profil"
            pie1.height, pie1.width = 6, 11
            data = Reference(ws, min_col=COL_VAL, min_row=5, max_row=r_fin)
            cats = Reference(ws, min_col=COL_LAB, min_row=5, max_row=r_fin)
            pie1.add_data(data, titles_from_data=False)
            pie1.set_categories(cats)
            pie1.dataLabels = _labels_cat_pourcentage()
            # Tracer les données même si les colonnes source sont masquées
            # (sinon Excel n'affiche rien quand V/W sont cachées).
            pie1.visible_cells_only = False
            ws.add_chart(pie1, "E4")
            # Colonnes techniques (source du camembert) masquées : usage interne
            ws.column_dimensions[get_column_letter(COL_LAB)].hidden = True
            ws.column_dimensions[get_column_letter(COL_VAL)].hidden = True

        # Tableau 2 : par collaborateur
        ws.cell(row=12, column=1, value="Collaborateur")
        ws.cell(row=12, column=2, value="Profil")
        ws.cell(row=12, column=3, value="Consommation")
        ws.cell(row=12, column=4, value="%")
        _style_entete(ws, "A12:D12")
        indiv = []
        for nom, cdata in collab.items():
            t = sum(rr[-1] for rr in cdata["rows"]
                    if str(rr[2]).upper() == code)
            if t > 0:
                indiv.append((nom, cdata["role"], round(t, 3)))
        total_indiv = sum(t for _, _, t in indiv)
        row_i = 13
        for nom, role, t in indiv:
            ws.cell(row=row_i, column=1, value=nom)
            ws.cell(row=row_i, column=2, value=role)
            ws.cell(row=row_i, column=3, value=t)
            pct = ws.cell(row=row_i, column=4,
                          value=(t / total_indiv) if total_indiv else 0)
            pct.number_format = "0.0%"
            row_i += 1
        if indiv:
            ws.cell(row=row_i, column=1, value="TOTAL INDIVIDUEL").font = FONT_BOLD
            ws.cell(row=row_i, column=3,
                    value=round(total_indiv, 3)).font = FONT_BOLD
            pct_tot2 = ws.cell(row=row_i, column=4, value=1 if total_indiv else 0)
            pct_tot2.number_format = "0.0%"
            pct_tot2.font = FONT_BOLD
            for c in range(1, 5):
                ws.cell(row=row_i, column=c).fill = FILL_TOTAL
            _bordures(ws, 12, 1, row_i, 4)

            # Camembert 2 : par collaborateur
            pie2 = PieChart()
            pie2.title = "Part par Collaborateur"
            pie2.height, pie2.width = 6.5, 11
            data2 = Reference(ws, min_col=3, min_row=13, max_row=row_i - 1)
            cats2 = Reference(ws, min_col=1, min_row=13, max_row=row_i - 1)
            pie2.add_data(data2, titles_from_data=False)
            pie2.set_categories(cats2)
            pie2.dataLabels = _labels_cat_pourcentage()
            ws.add_chart(pie2, "E22")

    # Ajustement automatique de la largeur des colonnes sur TOUS les onglets
    for ws in wb.worksheets:
        _ajuster_colonnes(ws)

    # Placer l'onglet Stats en première position
    if "Stats" in wb.sheetnames:
        wb._sheets.remove(ws_stats)
        wb._sheets.insert(0, ws_stats)
        wb.active = 0  # onglet actif à l'ouverture = Stats

    wb.save(sortie_path)


# =====================================================================
# MAIN
# =====================================================================
def chrono(nom, fn):
    t0 = time.perf_counter()
    res = fn()
    print(f"   ⏱ {nom} : {time.perf_counter() - t0:.2f}s")
    return res


def ouvrir_fichier(chemin):
    """Ouvre un fichier avec l'application par défaut du système.
    Multiplateforme (macOS / Windows / Linux). N'interrompt jamais le script :
    si l'ouverture échoue, le fichier reste généré et accessible manuellement.
    """
    try:
        if sys.platform == "darwin":          # macOS
            subprocess.run(["open", chemin], check=False)
        elif sys.platform.startswith("win"):  # Windows
            os.startfile(chemin)              # type: ignore[attr-defined]
        else:                                  # Linux et autres
            subprocess.run(["xdg-open", chemin], check=False)
    except Exception as e:
        print(f"   (Ouverture automatique impossible : {e})")
        print(f"   Ouvre le fichier manuellement : {chemin}")


def main():
    cfg = charger_config()
    token = cfg["jira"]["api_token"]
    # Placeholders connus (config exemple / questionnaire) = token non renseigné.
    placeholders = {"REMPLACE_PAR_TON_TOKEN", "REMPLACE_PAR_TON_NOUVEAU_TOKEN",
                    "colle_ton_token_ici"}
    if not token or token in placeholders:
        print(f"❌ Token Jira manquant dans {SECRETS_PATH}")
        return

    dict_param = {str(n).strip().upper(): str(r).strip()
                  for n, r in cfg.get("ressources", {}).items()}
    # normpath : tolère les séparateurs / ou \ quelle que soit la plateforme
    csv_path = os.path.normpath(cfg["chemins"]["csv"])

    if not os.path.exists(csv_path):
        print(f"❌ CSV introuvable : {csv_path}")
        print("   Vérifie le chemin 'csv' dans config.toml.")
        return

    # Dossier de sortie (depuis la config), créé s'il n'existe pas.
    dossier_sortie = os.path.normpath(cfg["chemins"]["dossier_sortie"])
    os.makedirs(dossier_sortie, exist_ok=True)

    # Nom du fichier généré : SPII_vs_SP_<PROJET>_<date>_<heure>.xlsx
    # Le projet vient de la config ; on retire d'éventuels caractères interdits
    # dans un nom de fichier (/ \ : * ? " < > |).
    projet = str(cfg["jira"].get("projet", "")).strip()
    projet_safe = re.sub(r'[\\/:*?"<>|]', "_", projet)
    horodatage = datetime.now().strftime("%Y-%m-%d_%Hh%M")
    bloc_projet = f"{projet_safe}_" if projet_safe else ""
    sortie = os.path.join(dossier_sortie,
                          f"SPII_vs_SP_{bloc_projet}{horodatage}.xlsx")

    # Préfixe des features (ex. "TCRE"), configurable.
    prefixe = str(cfg["jira"].get("prefixe_feature", "TCRE")).strip() or "TCRE"

    print("Construction du modèle métier...")
    modele = chrono("Lecture CSV + modèle",
                    lambda: construire_modele(csv_path, dict_param, prefixe))

    codes = list(modele["features_conso"].keys())
    print(f"Appels Jira ({len(codes)} {prefixe}, parallèle)...")
    jira = chrono("Jira", lambda: recuperer_jira(codes, cfg))

    print(f"Écriture du classeur openpyxl -> {os.path.basename(sortie)}...")
    chrono("Écriture openpyxl",
           lambda: ecrire_classeur(modele, jira, sortie, cfg))

    print(f"\n✅ Terminé : {sortie}")

    # Ouverture automatique du fichier généré
    ouvrir_fichier(sortie)


if __name__ == "__main__":
    main()